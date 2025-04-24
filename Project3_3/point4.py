import sys
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton, 
    QFileDialog, QMessageBox, QHBoxLayout, QScrollArea, QInputDialog,
    QLabel, QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView
)
from PyQt6.QtCore import Qt, QSize
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


class ExcelCellMarkerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Cell Marker")
        self.setGeometry(100, 100, 1000, 700)  # Increased window size
        
        # Variables to store workbook and cell selections
        self.workbook = None
        self.file_path = ""
        self.cell_selections = []  # List of selector dictionaries
        self.current_sheet_name = ""
        self.table_widget = None
        
        # Main widgets
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)
        
        # File selection button
        self.btn_select_file = QPushButton("Select Excel File")
        self.btn_select_file.clicked.connect(self.select_excel_file)
        self.layout.addWidget(self.btn_select_file)
        
        # Current file label
        self.lbl_current_file = QLabel("No file selected")
        self.layout.addWidget(self.lbl_current_file)
        
        # Excel table display
        self.table_widget = QTableWidget()
        self.table_widget.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_widget.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectItems)
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_widget.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.layout.addWidget(self.table_widget)
        
        # Scroll area for cell selection buttons
        self.scroll_area = QScrollArea()
        self.scroll_content = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_content)
        self.scroll_area.setWidget(self.scroll_content)
        self.scroll_area.setWidgetResizable(True)
        self.layout.addWidget(self.scroll_area)
        
        # Button to add more cell selectors
        self.btn_add_selector = QPushButton("Add Cell Selector")
        self.btn_add_selector.clicked.connect(self.add_cell_selector)
        self.btn_add_selector.setEnabled(False)
        self.layout.addWidget(self.btn_add_selector)
        
        # Save button
        self.btn_save = QPushButton("Save Changes to Excel")
        self.btn_save.clicked.connect(self.save_changes)
        self.btn_save.setEnabled(False)
        self.layout.addWidget(self.btn_save)
        
        # Add one initial selector
        self.add_cell_selector()
    
    def select_excel_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            
            self, "Select Excel File", "", "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            try:
                self.file_path = file_path
                self.workbook = openpyxl.load_workbook(self.file_path) 
                self.lbl_current_file.setText(f"Current file: {file_path}")
                self.btn_add_selector.setEnabled(True)
                self.btn_save.setEnabled(True)
                
                # Load the first sheet by default
                self.load_sheet(self.workbook.active)
                
                QMessageBox.information(
                    self, "Success", f"Loaded Excel file: {file_path}\n\n"
                    "Click on cells in the table to select them, then click the 'Select Cell' buttons to assign them."
                )
            except Exception as e:
                QMessageBox.critical(
                    self, "Error", f"Failed to load Excel file: {str(e)}"
                )
    
    def load_sheet(self, sheet):
        """Load the given sheet into the table widget"""
        self.current_sheet_name = sheet.title
        self.table_widget.clear()
        
        # Get sheet dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Set table dimensions
        self.table_widget.setRowCount(max_row)
        self.table_widget.setColumnCount(max_col)
        
        # Set headers
        headers = [get_column_letter(i+1) for i in range(max_col)]
        self.table_widget.setHorizontalHeaderLabels(headers)
        self.table_widget.setVerticalHeaderLabels([str(i+1) for i in range(max_row)])
        
        # Populate table with cell values
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                item = QTableWidgetItem(str(cell.value) if cell.value is not None else "")
                self.table_widget.setItem(row-1, col-1, item)
    
    def add_cell_selector(self):
        selector_index = len(self.cell_selections) + 1
        selector_widget = QWidget()
        selector_layout = QHBoxLayout(selector_widget)
        
        # Button to select cell
        btn_select = QPushButton(f"Select Cell {selector_index}")
        btn_select.clicked.connect(lambda: self.assign_selected_cell(selector_index))
        
        # Label to show selected cell reference
        lbl_cell_ref = QLabel("Not selected")
        lbl_cell_ref.setStyleSheet("font-weight: bold; min-width: 80px;")
        
        # Label to show cell value
        lbl_cell_value = QLabel("No value")
        lbl_cell_value.setStyleSheet("font-style: italic; min-width: 200px;")
        
        # Button to remove this selector
        btn_remove = QPushButton("Remove")
        btn_remove.clicked.connect(lambda: self.remove_selector(selector_index))
        
        selector_layout.addWidget(btn_select)
        selector_layout.addWidget(lbl_cell_ref)
        selector_layout.addWidget(QLabel("Value:"))
        selector_layout.addWidget(lbl_cell_value)
        selector_layout.addWidget(btn_remove)
        
        self.scroll_layout.addWidget(selector_widget)
        
        # Store the reference
        self.cell_selections.append({
            "widget": selector_widget,
            "select_button": btn_select,
            "cell_ref_label": lbl_cell_ref,
            "cell_value_label": lbl_cell_value,
            "cell_ref": None,
            "cell_value": None
        })
    
    def assign_selected_cell(self, selector_index):
        """Assign the currently selected cell in the table to this selector"""
        selected_items = self.table_widget.selectedItems()
        
        if not selected_items:
            QMessageBox.warning(self, "Warning", "Please select a cell in the Excel table first")
            return
            
        selected_item = selected_items[0]
        row = selected_item.row() + 1
        col = selected_item.column() + 1
        cell_ref = f"{get_column_letter(col)}{row}"
        
        # Get the cell value from the workbook (not the table widget, to be sure)
        sheet = self.workbook[self.current_sheet_name]
        cell = sheet[cell_ref]
        cell_value = str(cell.value) if cell.value is not None else "Empty"
        
        # Update the GUI
        self.cell_selections[selector_index-1]["cell_ref"] = cell_ref
        self.cell_selections[selector_index-1]["cell_value"] = cell_value
        self.cell_selections[selector_index-1]["cell_ref_label"].setText(cell_ref)
        self.cell_selections[selector_index-1]["cell_value_label"].setText(cell_value)
    
    def remove_selector(self, selector_index):
        # Find the widget to remove
        for i, selector in enumerate(self.cell_selections):
            if i == selector_index - 1:
                selector["widget"].setParent(None)
                self.cell_selections.remove(selector)
                break
        
        # Renumber remaining selectors
        for i, selector in enumerate(self.cell_selections):
            selector["select_button"].setText(f"Select Cell {i+1}")
    
    def save_changes(self):
        if not self.workbook or not self.file_path:
            QMessageBox.warning(self, "Warning", "No Excel file loaded")
            return
            
        # Get all valid cell references
        cells_to_mark = [
            selector["cell_ref"] for selector in self.cell_selections 
            if selector["cell_ref"]
        ]
        
        if not cells_to_mark:
            QMessageBox.warning(self, "Warning", "No cells selected for marking")
            return
            
        try:
            sheet = self.workbook[self.current_sheet_name]  # Get active sheet
            
            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply borders to selected cells
            for cell_ref in cells_to_mark:
                cell = sheet[cell_ref]
                cell.border = thin_border
            
            # Save the workbook
            self.workbook.save(self.file_path)
            QMessageBox.information(
                self, "Success", f"Borders added to cells and file saved: {self.file_path}"
            )
            
            # Reload the sheet to show changes
            self.load_sheet(sheet)
        except Exception as e:
            QMessageBox.critical(
                self, "Error", f"Failed to save changes: {str(e)}"
            )


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelCellMarkerApp()
    window.show()
    sys.exit(app.exec())